import xlrd

from openpyxl import workbook as wbk
from openpyxl.cell import *
import datetime
from easygui import fileopenbox,msgbox,choicebox
# from tempfile import TemporaryFile

def MySort(info):
    return int(info.split("$")[0])


fname = fileopenbox("Выберите файл с днями рождения", "")
if not fname:
    exit()
if not fname.endswith(".xlsx"):
        msgbox("Выбран не файл xlsx", ok_button="Закрыть", title="Проверьте тип файла!")
        exit()
rb = xlrd.open_workbook(fname)
all_persons = list()


sheets_list = rb.sheet_names() #нашли количество листов в книге
choice_list=[]

for c in sheets_list:
    choice_list.append(str(sheets_list.index(c)+1)+"."+str(c))

selected_sheet=choicebox(msg="Выберите лист для обработки",title="Ваш выбор",choices=choice_list)

# ws=rb.sheet_by_index(len(sheets_list)-1)
sheet=rb.sheet_by_name(selected_sheet.split(".")[1])
old_sheet_name= sheet.name
# print(dirname)
for rownum in range(sheet.nrows):

    rows=sheet.row_values(rownum)
    if sheet.cell(rownum,1).ctype !=3:
        continue
    date_tuple = xlrd.xldate_as_tuple(rows[1],rb.datemode)
    full_date =  datetime.datetime(*date_tuple)

     #єто день рождения (именно ДЕНЬ)
    day_of_birth = date_tuple[2]
    year_of_birth=date_tuple[0]
    info = str(day_of_birth)+ "$"+rows[0] + "$" + rows[2]+ "$" +str(year_of_birth) # сформировали строку из ФИО и адреса

    try:

        all_persons.append(info)

    except AttributeError:
        print("Ошибка добавления в словарь!")

    info=None

all_persons.sort(key=MySort)

new_wb=wbk.Workbook()
#new_wb=Workbook(encoding="utf-8")
del sheet
new_wb.remove_sheet(new_wb.active)

sheet=new_wb.create_sheet()
sheet.title="Сортировано - "+old_sheet_name


sheet.cell(row=1,column=1,value="День рождения")
sheet.cell(row=1,column=2,value="Год рождения")
sheet.cell(row=1,column=3,value="ФИО")
sheet.cell(row=1,column=4,value="Адрес")


row_count=2
for info in all_persons:
    bd,fio,adr,year_of_birth=info.split("$")
    sheet.cell(row=row_count,column=1,value=int(bd))
    sheet.cell(row=row_count,column=2, value=int(year_of_birth))
    sheet.cell(row=row_count,column=3,value=fio)
    sheet.cell(row=row_count,column=4,value=adr)


    row_count=row_count+1

new_wb.save(filename="Sorted-"+old_sheet_name+".xlsx")

pass
msgbox("Конец!")
exit()